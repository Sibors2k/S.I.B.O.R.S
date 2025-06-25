# src/utils/excel_reporter.py

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_reporte_excel(ruta_archivo: str, titulo: str, headers: list, data: list):
    """
    Genera un reporte en formato .xlsx.

    Args:
        ruta_archivo (str): Ruta completa donde se guardará el archivo.
        titulo (str): El título que aparecerá en la primera fila.
        headers (list): Una lista con los nombres de las columnas.
        data (list): Una lista de listas, donde cada lista interna es una fila de datos.
    """
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte"

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Título del Reporte ---
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1, value=titulo)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center_align

    # --- Encabezados de la Tabla ---
    sheet.append(headers) # Añade los encabezados en la siguiente fila disponible (fila 2)
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Datos de la Tabla ---
    for row_data in data:
        sheet.append(row_data)

    # --- Ajuste de Ancho de Columnas ---
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].autosize = True

    # Guardar el libro de trabajo
    workbook.save(ruta_archivo)
    print(f"📄 Reporte Excel generado en: {ruta_archivo}")